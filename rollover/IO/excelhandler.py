'''
Created on May 18, 2017

@author: kevin
'''

from .iohandler import IOHandler
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
import os.path

class ExcelHandler(IOHandler):      
    def __init__(self, abs_path):
        super().__init__(abs_path)
          
    def ReadSingleFile(self, parent = None):
        read_data = {} #{contract_name:data}
        wb = load_workbook(self.absPath, data_only=True, read_only=True) 
        wb_name = self.GetWorkBookNameFromAbsolutePath(self.absPath)
        for each_sheet in wb.worksheets:
            contract_name = wb_name + "." + each_sheet.title
            contract_data = []
            for row in each_sheet.iter_rows():
                if not isinstance(row[0].value, datetime): continue
                new_data = [cell.value for cell in row[:6]]
                contract_data.append(new_data)
            read_data[contract_name] = contract_data
        return read_data
        
    def CountSheetInOneExcel(self, abs_path):    
        totalsheet = 0
        wb = load_workbook(abs_path, data_only=True, read_only=True) 
        totalsheet += len(wb.worksheets)
        return totalsheet
            
    def GetWorkBookNameFromAbsolutePath(self, absolute_path):
        if absolute_path[-5:] == ".xlsx":
            return os.path.basename(absolute_path)[:-5]
        elif absolute_path[-4:] == ".xls":
            return os.path.basename(absolute_path)[:-4]      
        
    def SaveFile(self, data): 
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "time"
        ws["B1"] = "open"
        ws["C1"] = "high"
        ws["D1"] = "low"
        ws["E1"] = "close"
        ws["F1"] = "volume" 
        ws["G1"] = "adjusment" 
        ws["H1"] = "contract name"         
#         

        
        for i in range(len(data)):
            ws.append(data[i])
#             progressDialog.setValue(i)
        wb.save(self.absPath)
#         progressDialog.setValue(len(data))